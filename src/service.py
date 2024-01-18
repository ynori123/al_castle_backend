from openpyxl import load_workbook
from src.model import (
    Castle,
    Restaurant
)
from src.database import SessionLocal
from logging import getLogger
logger = getLogger("uvicorn.app")
from src.google_api import search_castle_address

def load_data() -> tuple[list[Castle], list[Restaurant]]:
    castle_data = []
    restaurant_data = []
    with SessionLocal() as db:
        logger.info("start load database")
        if db.query(Castle).count() != 0 and db.query(Restaurant).count() != 0:
            logger.info("data already exist.")
            return
    try:
        wb = load_workbook("castle_data.xlsx")
        print("start load data")
        # 飲食店データの読み込み
        ws = wb["castle_data"]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            
            if isinstance(row[0].value, int) == False:
                # 1列目が数字でない場合はスキップ
                logger.info(row[0].value)
                continue
            castle = Castle(
                name=row[1].value,
                prefecture=row[2].value,
                lat=float(row[3].value),
                lng=float(row[4].value),
                holiday=row[5].value,
                admission_time=row[6].value,
                admission_fee=row[7].value,
                stamp=row[8].value,
                address=search_castle_address(row[1].value)
            )
            castle_data.append(castle)
        logger.info("finish load castle data")
        logger.info("start load meal data")
        ws = wb["meal"]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            if isinstance(row[0].value, int) == False or isinstance(row[1].value, int) == False:
                # 1列目が数字でない場合はスキップ
                logger.info(f"load-error: {row[0].value}, type: {type(row[0].value)} / {row[1].value}, type: {type(row[1].value)}")
                continue
            restaurant = Restaurant(
                castle_id=int(row[1].value),
                name=row[2].value,
                time=row[3].value,
                holiday=row[4].value,
                genre=row[5].value,
                url=row[6].value,
            )
            restaurant_data.append(restaurant)
        logger.info("finish load meal data")
        return (castle_data, restaurant_data)
    except Exception as e:
        logger.info(f"error: {e}")
        return 

def write_data(castle_data: list[Castle], restaurant_data: list[Restaurant]):
    with SessionLocal() as db:
        try:
            if db.query(Castle).count() == 0:
                db.add_all(castle_data)
            if db.query(Restaurant).count() == 0:
                valid_restaurant_data = [r for r in restaurant_data if r.name and r.genre is not None]
                db.add_all(valid_restaurant_data)
            db.commit()
        except Exception as e:
            db.rollback()
            raise e
    return
